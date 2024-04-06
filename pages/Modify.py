import streamlit as st
import pandas as pd
from openpyxl import load_workbook

def main():
    already = dict()

    # Load the Excel file into a DataFrame
    excel_file_path = 'data\\customer_buy.xlsx'  # Replace with your file path
    df_products = pd.read_excel(excel_file_path)

    if(df_products.shape[0] == 0):
        st.warning("Select some products")
    else:
        st.title(":blue[Product List]")
        for index, product in df_products.iterrows():
            ori = f"{product['product_name']}"
            name = f"{product['product_name']}"
            if ori in already:
                name = f"{product['product_name']}({already[ori]})"
                already[ori] += 1
            else:
                already[ori] = 1

            st.subheader(name)
            
            # Delete button for each product
            delete_product = st.button(name)
            
            if delete_product:
                df_products.drop(index, inplace=True)
                df_products.to_excel(excel_file_path, index=False)
                st.experimental_rerun()
            
            st.write("---")

        # Clear button to delete all products
        clear_all = st.button("Clear All Products")
        
        if clear_all:
            wb = load_workbook(excel_file_path)
            ws = wb.active
            ws.delete_rows(2, ws.max_row)  # Delete all rows except the header
            wb.save(excel_file_path)
            st.write("All products cleared!")
            st.experimental_rerun()

if __name__ == "__main__":
    main()
